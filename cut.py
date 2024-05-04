import os
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor

def list_excel_files(directory):
    excel_files = []
    for file in os.listdir(directory):
        if file.endswith(".xlsx"):
            excel_files.append(file)
    return excel_files

def create_redacted_directory(directory):
    redacted_directory = os.path.join(directory, "redacted")
    if not os.path.exists(redacted_directory):
        os.makedirs(redacted_directory)

def trim_excel_single(file_path, redacted_directory):
    wb = load_workbook(file_path)
    ws = wb.active
    rows_to_delete = []
    for row in ws.iter_rows(min_row=2):
        if row[4].value is None:
            rows_to_delete.append(row)
    for row in rows_to_delete:
        ws.delete_rows(row[0].row)
    new_file_name = os.path.basename(file_path).replace(".xlsx", "_redacted.xlsx")
    new_file_path = os.path.join(redacted_directory, new_file_name)
    wb.save(new_file_path)
    return new_file_path

def trim_excel(directory, file, redacted_directory):
    file_path = os.path.join(directory, file)
    return trim_excel_single(file_path, redacted_directory)

def trim_excel_multiple(directory):
    create_redacted_directory(directory)
    excel_files = list_excel_files(directory)
    redacted_directory = os.path.join(directory, "redacted")
    with ThreadPoolExecutor(max_workers=8) as executor:
        results = executor.map(lambda f: trim_excel(directory, f, redacted_directory), excel_files)
    return list(results)

def main():
    directory = "./db"
    print("*=====================*")
    print("Database Trimmer\n> This means that all cells in the database that do not contain a phone number will be deleted.\nThe file is trimmed in 8 threads, you can change the number of threads in the code yourself.")
    print("*=====================*\n")
    option = input("[1] Trim a single database\n[2] Trim all databases\nChoose an option: ")
    if option == '1':
        excel_files = list_excel_files(directory)
        print("Choose the database number:")
        for i, file in enumerate(excel_files):
            print(f"[{i+1}] {file}")
        chosen_file = int(input(">> "))
        print("Trimming...")
        new_file_name = trim_excel_single(os.path.join(directory, excel_files[chosen_file-1]), os.path.join(directory, "redacted"))
        print(f"Done. {new_file_name} is located in the folder {directory}/redacted")
    elif option == '2':
        print("Trimming...")
        new_files = trim_excel_multiple(directory)
        for file in new_files:
            print(f"Done. {file} is located in the folder {directory}/redacted")
    else:
        print("Incorrect option choice.")

if __name__ == "__main__":
    main()