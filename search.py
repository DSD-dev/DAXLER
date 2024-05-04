import os
from openpyxl import load_workbook
from termcolor import colored

def list_excel_files(directory):
    excel_files = []
    for file in os.listdir(directory):
        if file.endswith(".xlsx"):
            excel_files.append(file)
    return excel_files

def choose_option():
    print("[1] Search in a single database")
    print("[2] Search in all databases")
    choice = input("Choose an option: ")
    return choice

def choose_excel_file(excel_files):
    print("Choose the database number:")
    for i, file in enumerate(excel_files):
        print(f"[{i+1}] {file}")
    choice = int(input(">> "))
    return excel_files[choice-1]

def search_excel(file_path, query):
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(query) in str(row[0]):
            return (row[0], row[2], row[4], row[6], row[8])
    return None

def search_all_excel_files(directory, query):
    excel_files = list_excel_files(directory)
    for file in excel_files:
        print(colored(f"Searching for '{query}' in {file}...", "blue"))
        result = search_excel(os.path.join(directory, file), query)
        if result:
            print(f"{colored('ID:', 'blue')} {result[0]}")
            print(f"{colored('NAME:', 'blue')} {result[1]}")
            print(f"{colored('NUMBER:', 'blue')} {result[2]}")
            print(f"{colored('USERNAME:', 'blue')} {result[3]}")
            print(f"{colored('LINK:', 'blue')} {result[4]}")
        else:
            print(colored("Nothing found.", "red"))
        print()

def main():
    directory = "./db"  # Directory where Excel files are located
    option = choose_option()
    if option == '1':
        excel_files = list_excel_files(directory)
        chosen_file = choose_excel_file(excel_files)
        query = input("Enter a number to search for: ")
        print(colored(f"Searching for '{query}' in {chosen_file}...", "blue"))
        result = search_excel(os.path.join(directory, chosen_file), query)
        if result:
            print(f"{colored('ID:', 'blue')} {result[0]}")
            print(f"{colored('NAME:', 'blue')} {result[1]}")
            print(f"{colored('NUMBER:', 'blue')} {result[2]}")
            print(f"{colored('USERNAME:', 'blue')} {result[3]}")
            print(f"{colored('LINK:', 'blue')} {result[4]}")
        else:
            print(colored("Nothing found.", "red"))
    elif option == '2':
        query = input("Enter a query to search for: ")
        search_all_excel_files(directory, query)
    else:
        print(colored("Invalid option choice.", "red"))

if __name__ == "__main__":
    main()